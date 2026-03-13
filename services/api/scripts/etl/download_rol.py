"""
ETL: Download e ingestão do Rol de Procedimentos (Anexo I) da ANS.

Fonte: Excel oficial publicado no gov.br.
"""
import asyncio
import hashlib
import io
from datetime import datetime
from pathlib import Path

import httpx
from sqlalchemy import select

ROL_XLSX_URL = "https://www.gov.br/ans/pt-br/acesso-a-informacao/participacao-da-sociedade/atualizacao-do-rol-de-procedimentos/Anexo_I_Rol_2021RN_465.2021_RN643.2025.xlsx/@@download/file"
DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "ans"


def compute_sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


async def download_rol_xlsx(url: str = ROL_XLSX_URL) -> bytes:
    async with httpx.AsyncClient(follow_redirects=True, timeout=120) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        return resp.content


def parse_rol_xlsx(xlsx_bytes: bytes) -> list[dict]:
    """Parseia o Anexo I (Excel) e retorna lista de procedimentos."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        # Find the actual header row: must have "PROCEDIMENTO" as an exact cell value
        # (not as part of a title like "Rol de Procedimentos e Eventos em Saúde")
        header_row_idx = 0
        header = []
        for ri, r in enumerate(rows):
            cells = [str(c).strip().upper() if c else "" for c in r]
            cell_values = {c for c in cells if c}
            if "PROCEDIMENTO" in cell_values or "CÓDIGO" in cell_values or "CODIGO" in cell_values:
                header = cells
                header_row_idx = ri
                break

        if not header:
            continue

        col_map = {}
        for i, h in enumerate(header):
            h_lower = h.lower()
            if "procedimento" in h_lower or "código" in h_lower or "codigo" in h_lower:
                if "código" in h_lower or "codigo" in h_lower:
                    col_map["codigo"] = i
                elif "procedimento" in h_lower:
                    col_map.setdefault("nome", i)
            elif h_lower in ("amb", "ambulat") or "ambulat" in h_lower:
                col_map["amb"] = i
            elif h_lower in ("hco",) or ("hospit" in h_lower and "obst" not in h_lower):
                col_map["hosp"] = i
            elif h_lower in ("hso",) or "obst" in h_lower:
                col_map["obst"] = i
            elif h_lower in ("od",) or "odonto" in h_lower:
                col_map["odonto"] = i
            elif "dut" in h_lower or "diretriz" in h_lower:
                col_map["dut"] = i
            elif "grupo" in h_lower and "sub" not in h_lower:
                col_map["grupo"] = i
            elif "subgrupo" in h_lower:
                col_map["subgrupo"] = i
            elif h_lower == "ref":
                col_map["ref"] = i

        # If no explicit "codigo" column, use "nome" as the procedure identifier
        if "codigo" not in col_map and "nome" in col_map:
            col_map["codigo"] = col_map["nome"]

        if "codigo" not in col_map:
            continue

        for row in rows[header_row_idx + 1:]:
            raw_code = row[col_map["codigo"]] if col_map["codigo"] < len(row) else None

            if not raw_code:
                continue

            code = str(raw_code).strip()
            if not code or code == "None":
                continue
            # Skip title/subtitle/header echo rows
            if code.startswith("(") or code.startswith("Rol ") or code == "PROCEDIMENTO":
                continue

            # Use the name column if separate from code, otherwise code IS the name
            name = code
            if "nome" in col_map and col_map["nome"] != col_map["codigo"]:
                raw_name = row[col_map["nome"]] if col_map["nome"] < len(row) else None
                name = str(raw_name).strip() if raw_name else code

            def _seg_bool(col_key):
                idx = col_map.get(col_key)
                if idx is None or idx >= len(row):
                    return False
                v = row[idx]
                if v is None:
                    return False
                sv = str(v).strip().upper()
                return sv in ("SIM", "S", "X", "1", "TRUE", "OAM", "AMB", "HCO", "HSO", "OD", "REF")

            # DUT detection: explicit column OR "(COM DIRETRIZ DE UTILIZAÇÃO)" in name
            dut_val = ""
            if "dut" in col_map and col_map["dut"] < len(row):
                dut_val = str(row[col_map["dut"]] or "").strip()

            has_dut = bool(dut_val and dut_val.lower() not in ("", "none", "nan", "-"))
            if not has_dut and "diretriz de utilização" in code.lower():
                has_dut = True
                dut_val = "sim"

            raw_data = {}
            for ci, cv in enumerate(row):
                if cv is not None and ci < len(header):
                    raw_data[header[ci]] = str(cv)

            records.append({
                "codigo_procedimento": code[:50],
                "nome": name,
                "segmentacao_ambulatorial": _seg_bool("amb"),
                "segmentacao_hospitalar": _seg_bool("hosp"),
                "segmentacao_obstetrica": _seg_bool("obst"),
                "segmentacao_odontologica": _seg_bool("odonto"),
                "tem_dut": has_dut,
                "dut_numero": dut_val if has_dut else None,
                "grupo": str(row[col_map["grupo"]]).strip() if "grupo" in col_map and col_map["grupo"] < len(row) and row[col_map["grupo"]] else None,
                "subgrupo": str(row[col_map["subgrupo"]]).strip() if "subgrupo" in col_map and col_map["subgrupo"] < len(row) and row[col_map["subgrupo"]] else None,
                "raw_data": raw_data,
            })

    wb.close()
    return records


async def ingest_rol_procedures(db_session, records: list[dict], version_id) -> dict:
    """Upsert de procedimentos do Rol no banco."""
    from app.db.models import RolProcedure

    inserted = 0
    updated = 0
    for rec in records:
        existing = await db_session.execute(
            select(RolProcedure).where(
                RolProcedure.codigo_procedimento == rec["codigo_procedimento"],
                RolProcedure.version_id == version_id,
            )
        )
        existing = existing.scalar_one_or_none()
        if existing:
            for k, v in rec.items():
                if k != "raw_data":
                    setattr(existing, k, v)
            existing.raw_data = rec["raw_data"]
            updated += 1
        else:
            proc = RolProcedure(**rec, version_id=version_id)
            db_session.add(proc)
            inserted += 1

    await db_session.commit()
    return {"inserted": inserted, "updated": updated, "total": len(records)}


async def run_etl(db_session=None):
    """Executa ETL completo do Rol."""
    print("[ROL ETL] Baixando Anexo I...")
    xlsx_bytes = await download_rol_xlsx()
    sha = compute_sha256(xlsx_bytes)
    print(f"[ROL ETL] XLSX baixado: {len(xlsx_bytes)} bytes, SHA256: {sha[:16]}...")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    (DATA_DIR / "Rol_Anexo_I.xlsx").write_bytes(xlsx_bytes)

    print("[ROL ETL] Parseando procedimentos...")
    records = parse_rol_xlsx(xlsx_bytes)
    print(f"[ROL ETL] {len(records)} procedimentos encontrados")

    if db_session:
        from app.db.models import RolVersion
        version = RolVersion(
            versao=datetime.utcnow().strftime("%Y%m%d"),
            rn_numeros=["465/2021", "643/2025"],
            hash_arquivo=sha,
            url_fonte=ROL_XLSX_URL,
            data_publicacao=datetime.utcnow(),
        )
        db_session.add(version)
        await db_session.flush()
        result = await ingest_rol_procedures(db_session, records, version.id)
        print(f"[ROL ETL] Resultado: {result}")
        return result

    return {"records_parsed": len(records), "sha256": sha}


if __name__ == "__main__":
    asyncio.run(run_etl())
