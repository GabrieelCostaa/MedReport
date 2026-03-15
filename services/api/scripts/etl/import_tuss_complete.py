"""
ETL: Importa Tabela TUSS 19 COMPLETA (Materiais e OPME) do pacote TISS da ANS.

A Tabela 19 completa vem do "Padrão TISS - Representação de Conceitos em Saúde",
NÃO do portal de dados abertos (que só tem 15 categorias).

Fonte: https://www.ans.gov.br/arquivos/extras/tiss/Padrao_TISS_Representacao_de_Conceitos_em_Saude_YYYYMM.zip
Contém 2 arquivos XLSX com 74.000+ códigos TUSS de materiais/OPME.
"""
import asyncio
import io
import os
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent.parent))

from scripts.etl.download_tuss import normalize_text

TISS_ZIP_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "ans" / "tiss" / "TISS_202601.zip"


def extract_table_19_from_tiss_zip(zip_path: Path) -> list[dict]:
    """Extract Table 19 XLSX files from the TISS ZIP and parse all records."""
    records = []

    with zipfile.ZipFile(zip_path) as zf:
        # Find Table 19 XLSX files
        xlsx_files = [
            n for n in zf.namelist()
            if n.endswith(".xlsx") and "19" in n and ("material" in n.lower() or "opme" in n.lower())
            and not os.path.basename(n).startswith("~$")  # Skip Excel temp files
        ]

        if not xlsx_files:
            print(f"[WARN] No Table 19 XLSX found. Files in ZIP:")
            for n in sorted(zf.namelist()):
                if n.endswith((".xlsx", ".csv")):
                    print(f"  {n}")
            return records

        print(f"[INFO] Found {len(xlsx_files)} Table 19 XLSX files:")
        for f in xlsx_files:
            print(f"  {f}")

        for xlsx_name in xlsx_files:
            print(f"\n[INFO] Parsing {xlsx_name}...")
            xlsx_bytes = zf.read(xlsx_name)
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
            ws = wb.active

            # Find header row (first row with "Código" or "CODIGO" or similar)
            header = None
            header_row_idx = 0
            for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                row_str = " ".join(str(c or "").lower() for c in row)
                if "digo" in row_str and ("termo" in row_str or "descri" in row_str):
                    header = [str(c or "").strip() for c in row]
                    header_row_idx = i + 1
                    break

            if not header:
                print(f"  [WARN] No header found in {xlsx_name}, trying first row")
                rows = list(ws.iter_rows(max_row=2, values_only=True))
                if rows:
                    header = [str(c or "").strip() for c in rows[0]]
                    header_row_idx = 1
                else:
                    continue

            print(f"  Header (row {header_row_idx}): {header[:6]}...")

            # Map column indices
            col_map = {}
            for idx, col_name in enumerate(header):
                cl = col_name.lower().strip()
                if "digo" in cl and "termo" in cl:
                    col_map["codigo"] = idx
                elif cl in ("termo", "descrição", "descricao"):
                    col_map["nome"] = idx
                elif "modelo" in cl:
                    col_map["modelo"] = idx
                elif "fabricante" in cl:
                    col_map["fabricante"] = idx
                elif "registro" in cl and "anvisa" in cl:
                    col_map["registro_anvisa"] = idx
                elif "classe" in cl and "risco" in cl:
                    col_map["classe_risco"] = idx
                elif "nome" in cl and ("cnico" in cl or "tecnico" in cl):
                    col_map["nome_tecnico"] = idx
                elif "in" in cl and "cio" in cl and "vig" in cl:
                    col_map["inicio_vigencia"] = idx
                elif "fim" in cl and "vig" in cl:
                    col_map["fim_vigencia"] = idx

            print(f"  Mapped columns: {col_map}")

            if "codigo" not in col_map or "nome" not in col_map:
                # Try alternate: first col = codigo, second = nome
                col_map["codigo"] = 0
                col_map["nome"] = 1
                print(f"  [WARN] Using fallback column mapping: col 0=codigo, col 1=nome")

            row_count = 0
            skip_count = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                code_val = row[col_map["codigo"]] if col_map["codigo"] < len(row) else None
                name_val = row[col_map["nome"]] if col_map["nome"] < len(row) else None

                code = str(code_val or "").strip()
                name = str(name_val or "").strip()

                if not code or not name or code.lower() in ("código do termo", "codigo"):
                    skip_count += 1
                    continue

                # Extract optional fields
                fabricante = ""
                if "fabricante" in col_map and col_map["fabricante"] < len(row):
                    fabricante = str(row[col_map["fabricante"]] or "").strip()

                reg_anvisa = ""
                if "registro_anvisa" in col_map and col_map["registro_anvisa"] < len(row):
                    reg_anvisa = str(row[col_map["registro_anvisa"]] or "").strip()

                nome_tecnico = ""
                if "nome_tecnico" in col_map and col_map["nome_tecnico"] < len(row):
                    nome_tecnico = str(row[col_map["nome_tecnico"]] or "").strip()

                records.append({
                    "codigo_tuss": code,
                    "nome": name,
                    "display_normalized": normalize_text(name),
                    "grupo": nome_tecnico or None,
                    "subgrupo": None,
                    "fabricante": fabricante or None,
                    "manufacturer_normalized": normalize_text(fabricante) if fabricante else None,
                    "registro_anvisa": reg_anvisa or None,
                    "descricao": None,
                    "ativo": True,
                    "raw_data": None,  # Skip raw_data to save space (74k+ rows)
                })
                row_count += 1

            wb.close()
            print(f"  Parsed {row_count:,} records (skipped {skip_count})")

    return records


async def ingest_records(records: list[dict], versao: str):
    """Bulk upsert records into tuss_materials table."""
    from app.db.session import AsyncSessionLocal
    from app.db.models import TussMaterial
    from sqlalchemy import select, text

    print(f"\n[INFO] Ingesting {len(records):,} records into tuss_materials...")

    async with AsyncSessionLocal() as db:
        # Get existing codes for fast lookup
        result = await db.execute(
            text("SELECT codigo_tuss FROM tuss_materials")
        )
        existing_codes = {row[0] for row in result.fetchall()}
        print(f"  Existing records: {len(existing_codes):,}")

        inserted = 0
        updated = 0
        batch_size = 500

        for i in range(0, len(records), batch_size):
            batch = records[i:i + batch_size]

            for rec in batch:
                if rec["codigo_tuss"] in existing_codes:
                    # Update existing
                    await db.execute(
                        text("""
                            UPDATE tuss_materials SET
                                nome = :nome,
                                display_normalized = :display_normalized,
                                grupo = COALESCE(:grupo, grupo),
                                fabricante = COALESCE(:fabricante, fabricante),
                                manufacturer_normalized = COALESCE(:manufacturer_normalized, manufacturer_normalized),
                                registro_anvisa = COALESCE(:registro_anvisa, registro_anvisa),
                                versao_tuss = :versao,
                                data_atualizacao = NOW()
                            WHERE codigo_tuss = :codigo_tuss
                        """),
                        {**rec, "versao": versao, "codigo_tuss": rec["codigo_tuss"]},
                    )
                    updated += 1
                else:
                    # Insert new
                    material = TussMaterial(
                        **rec,
                        versao_tuss=versao,
                        data_atualizacao=datetime.utcnow(),
                    )
                    db.add(material)
                    existing_codes.add(rec["codigo_tuss"])
                    inserted += 1

            await db.commit()

            if (i + batch_size) % 5000 == 0 or i + batch_size >= len(records):
                print(f"  Progress: {min(i + batch_size, len(records)):,}/{len(records):,} "
                      f"(inserted={inserted:,}, updated={updated:,})")

        print(f"\n[DONE] Total: inserted={inserted:,}, updated={updated:,}")

        # Stats
        result = await db.execute(text("SELECT count(*) FROM tuss_materials"))
        total = result.scalar()
        result = await db.execute(
            text("SELECT count(DISTINCT registro_anvisa) FROM tuss_materials WHERE registro_anvisa IS NOT NULL AND registro_anvisa != ''")
        )
        distinct_reg = result.scalar()
        result = await db.execute(
            text("SELECT count(DISTINCT grupo) FROM tuss_materials WHERE grupo IS NOT NULL")
        )
        distinct_groups = result.scalar()

        print(f"  DB total: {total:,} materials")
        print(f"  Distinct ANVISA registros: {distinct_reg:,}")
        print(f"  Distinct categories (grupo): {distinct_groups:,}")

        # Show new categories
        result = await db.execute(
            text("SELECT DISTINCT grupo, count(*) as cnt FROM tuss_materials WHERE grupo IS NOT NULL GROUP BY grupo ORDER BY cnt DESC LIMIT 30")
        )
        print(f"\n  Top 30 categories:")
        for row in result.fetchall():
            print(f"    {row[1]:>6,}  {(row[0] or '')[:70]}")


async def main():
    if not TISS_ZIP_PATH.exists():
        print(f"[ERROR] TISS ZIP not found at {TISS_ZIP_PATH}")
        print("Download it first:")
        print('  curl -L -o data/ans/tiss/TISS_202601.zip "https://www.ans.gov.br/arquivos/extras/tiss/Padrao_TISS_Representacao_de_Conceitos_em_Saude_202601.zip"')
        return

    print(f"[INFO] ZIP: {TISS_ZIP_PATH} ({TISS_ZIP_PATH.stat().st_size / 1024 / 1024:.1f} MB)")

    records = extract_table_19_from_tiss_zip(TISS_ZIP_PATH)
    print(f"\n[INFO] Total records parsed: {len(records):,}")

    if not records:
        print("[ERROR] No records parsed!")
        return

    # Show sample
    print(f"\nSample records:")
    for r in records[:3]:
        print(f"  {r['codigo_tuss']}: {r['nome'][:60]} | grupo={r['grupo'] or 'N/A'} | reg={r['registro_anvisa'] or 'N/A'}")

    # Distinct registros and categories
    regs = {r["registro_anvisa"] for r in records if r["registro_anvisa"]}
    cats = {r["grupo"] for r in records if r["grupo"]}
    print(f"\nDistinct ANVISA registros: {len(regs):,}")
    print(f"Distinct categories: {len(cats):,}")

    await ingest_records(records, "TISS_202601")


if __name__ == "__main__":
    asyncio.run(main())
