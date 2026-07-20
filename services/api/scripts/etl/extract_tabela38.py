"""Extração LOCAL (1x) da TISS Tabela 38 — Terminologia de mensagens (motivos de glosa).

A Tabela 38 NÃO vem no TUSS.zip dos dados abertos; ela é uma aba do XLSX
"TUSS - Demais terminologias" dentro do pacote do Padrão TISS (~559 MB).
Este script roda apenas em ambiente de desenvolvimento (o pacote não cabe
no Render free tier): extrai a aba e grava um CSV pequeno que é VERSIONADO
no repositório e ingerido em produção por scripts/etl/ingest_tabela38.py.

Uso:
    cd services/api && python scripts/etl/extract_tabela38.py

Para atualizar em versões futuras do TISS: baixe o pacote "Padrão TISS -
Componente de Representação de Conceitos em Saúde" em
https://www.gov.br/ans/pt-br/assuntos/prestadores/padrao-para-troca-de-informacao-de-saude-suplementar-2013-tiss
salve o zip em data/ans/tiss/ e rode novamente (ajuste TISS_ZIP_GLOB se preciso).
"""
import csv
import io
import sys
import zipfile
import unicodedata
from pathlib import Path

DATA_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "ans"
TISS_ZIP_GLOB = "tiss/TISS_*.zip"
OUTPUT_CSV = DATA_DIR / "tabela38_motivos_glosa.csv"
SHEET_NAME = "Tab 38"

# openpyxl em modo read_only reporta a DIMENSÃO declarada da planilha
# (~1,05M linhas nesta aba), não as linhas reais — por isso os limites abaixo.
MAX_EMPTY_ROWS = 50
MAX_DATA_ROWS = 20_000


def _strip_accents(text: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def find_terminologias_xlsx(namelist: list[str]) -> str | None:
    """Localiza o XLSX "Demais terminologias" no zip.

    O nome interno tem encoding quebrado ("VERS?O"/"VERSÃO" dependendo do
    unzip) — casamos por substring sem acentos e ignoramos locks do Office (~$).
    """
    for name in namelist:
        base = _strip_accents(name).lower()
        if "demais terminologias" in base and base.endswith(".xlsx") and "~$" not in name:
            return name
    return None


def parse_tabela38_sheet(xlsx_bytes: bytes, sheet_name: str = SHEET_NAME) -> list[dict]:
    """Parseia a aba Tab 38 → lista de dicts {codigo, descricao, vigencia_inicio, vigencia_fim}.

    Estrutura verificada (versão 202601): título na linha 4, header na linha 6
    ("Código do Termo", "Termo", "Data de início de vigência", "Data de fim de
    vigência", "Data de fim de implantação"), dados a partir da linha 7.
    O header é detectado dinamicamente nas primeiras 15 linhas para tolerar
    variações entre versões do TISS.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Aba '{sheet_name}' não encontrada. Abas: {wb.sheetnames[:10]}...")
    ws = wb[sheet_name]

    records: list[dict] = []
    header_found = False
    col_codigo = col_termo = col_vig_ini = col_vig_fim = None
    empty_streak = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]

        if not header_found:
            if row_idx > 15:
                raise ValueError("Header 'Código do Termo' não encontrado nas primeiras 15 linhas")
            lowered = [_strip_accents(c).lower() for c in cells]
            if any("codigo do termo" in c for c in lowered):
                for i, c in enumerate(lowered):
                    if "codigo do termo" in c:
                        col_codigo = i
                    elif c == "termo":
                        col_termo = i
                    elif "inicio de vigencia" in c:
                        col_vig_ini = i
                    elif "fim de vigencia" in c:
                        col_vig_fim = i
                if col_codigo is None or col_termo is None:
                    raise ValueError(f"Header incompleto na linha {row_idx}: {cells}")
                header_found = True
            continue

        codigo = cells[col_codigo] if col_codigo < len(cells) else ""
        termo = cells[col_termo] if col_termo < len(cells) else ""

        if not codigo and not termo:
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_ROWS:
                break
            continue
        empty_streak = 0

        if not codigo or not termo:
            continue  # linha parcial (ex.: nota de rodapé)

        def _cell(idx):
            if idx is None or idx >= len(cells):
                return ""
            val = cells[idx]
            # datetime vem como "2006-11-16 00:00:00" — mantém só a data
            return val.split(" ")[0] if val else ""

        records.append({
            "codigo": codigo,
            "descricao": termo,
            "vigencia_inicio": _cell(col_vig_ini),
            "vigencia_fim": _cell(col_vig_fim),
        })

        if len(records) >= MAX_DATA_ROWS:
            print(f"AVISO: teto de {MAX_DATA_ROWS} linhas atingido — verifique a aba", file=sys.stderr)
            break

    wb.close()
    return records


def write_csv(records: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(["codigo", "descricao", "vigencia_inicio", "vigencia_fim"])
        for r in records:
            writer.writerow([r["codigo"], r["descricao"], r["vigencia_inicio"], r["vigencia_fim"]])


def main() -> int:
    zips = sorted((DATA_DIR).glob(TISS_ZIP_GLOB))
    if not zips:
        print(
            f"Pacote TISS não encontrado em {DATA_DIR}/tiss/.\n"
            "Baixe o 'Padrão TISS - Componente de Representação de Conceitos em Saúde' em\n"
            "https://www.gov.br/ans/pt-br/assuntos/prestadores/padrao-para-troca-de-informacao-de-saude-suplementar-2013-tiss\n"
            "e salve como data/ans/tiss/TISS_<versao>.zip",
            file=sys.stderr,
        )
        return 1

    zip_path = zips[-1]  # versão mais recente
    print(f"Lendo {zip_path.name}...")
    with zipfile.ZipFile(zip_path) as zf:
        xlsx_name = find_terminologias_xlsx(zf.namelist())
        if not xlsx_name:
            print("XLSX 'Demais terminologias' não encontrado no zip", file=sys.stderr)
            return 1
        print(f"Extraindo {xlsx_name}...")
        xlsx_bytes = zf.read(xlsx_name)

    records = parse_tabela38_sheet(xlsx_bytes)
    if len(records) < 50:
        print(f"AVISO: apenas {len(records)} motivos extraídos — resultado suspeito", file=sys.stderr)

    write_csv(records, OUTPUT_CSV)
    print(f"OK: {len(records)} motivos de glosa gravados em {OUTPUT_CSV}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
